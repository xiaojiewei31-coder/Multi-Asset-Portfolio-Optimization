from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .config import ALTERNATIVE_PRICE_RANGES, ASSET_RANGES, DATA_PATH, NINE_ASSETS


def _to_month_period(value: object) -> pd.Period | None:
    if value is None:
        return None
    try:
        return pd.Period(pd.to_datetime(value), freq="M")
    except (TypeError, ValueError):
        try:
            return pd.Period(str(value), freq="M")
        except (TypeError, ValueError):
            return None


def load_monthly_returns(
    path: Path = DATA_PATH,
    sheet_name: str = "Sheet3",
    asset_ranges: dict[str, tuple[str, int, int]] = ASSET_RANGES,
) -> pd.DataFrame:
    """Load monthly returns from the project workbook."""
    workbook = load_workbook(filename=path, data_only=True)
    worksheet = workbook[sheet_name]

    returns = {}
    for asset, (column, start_row, end_row) in asset_ranges.items():
        values = [worksheet[f"{column}{row}"].value for row in range(start_row, end_row + 1)]
        returns[asset] = pd.to_numeric(pd.Series(values), errors="coerce")

    returns_df = pd.DataFrame(returns).dropna(how="any")
    returns_df.index = pd.period_range("2015-01", periods=len(returns_df), freq="M")
    return returns_df


def load_base_asset_returns(path: Path = DATA_PATH) -> pd.DataFrame:
    """Load the original five risky assets from the course workbook."""
    return load_monthly_returns(path)


def load_price_returns(
    path: Path = DATA_PATH,
    sheet_name: str = "Sheet2",
    price_ranges: dict[str, tuple[str, str, int, int]] = ALTERNATIVE_PRICE_RANGES,
) -> pd.DataFrame:
    """Load monthly prices and convert them to simple returns."""
    workbook = load_workbook(filename=path, data_only=True)
    worksheet = workbook[sheet_name]

    asset_prices = {}
    for asset, (date_column, price_column, start_row, end_row) in price_ranges.items():
        rows = []
        for row in range(start_row, end_row + 1):
            period = _to_month_period(worksheet[f"{date_column}{row}"].value)
            price = worksheet[f"{price_column}{row}"].value
            if period is None:
                continue
            rows.append((period, price))

        prices = pd.Series(
            data=[price for _, price in rows],
            index=[period for period, _ in rows],
            name=asset,
        )
        prices = pd.to_numeric(prices, errors="coerce").dropna()
        asset_prices[asset] = prices

    price_df = pd.DataFrame(asset_prices).sort_index()
    return price_df.pct_change().dropna(how="any")


def load_nine_asset_returns(path: Path = DATA_PATH) -> pd.DataFrame:
    """Combine the five original assets with four alternative asset classes."""
    base_returns = load_base_asset_returns(path)
    alternative_returns = load_price_returns(path)

    combined = pd.concat([alternative_returns, base_returns], axis=1, join="inner")
    return combined.loc[:, NINE_ASSETS].dropna(how="any")
